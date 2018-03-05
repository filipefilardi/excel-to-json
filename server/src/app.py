from flask import Flask, request, redirect, url_for, render_template, jsonify
from flask_pymongo import PyMongo

import os
import json
import glob

from openpyxl import Workbook , load_workbook
from openpyxl.utils import column_index_from_string

from config.keys import keys

app = Flask(__name__)

app.config['MONGO_DBNAME'] = keys['MONGO_DBNAME']
app.config['MONGO_URI'] = keys['MONGO_URI']

mongo = PyMongo(app)

def insert_data_db(json_data):
    try:
        converted = mongo.db.converted

        # already have one json with this name
        if (converted.find_one({'FILE_NAME': json_data["FILE_NAME"]})):
            return True
        # insert if do not have
        converted.insert(json_data)

        return True
    except:
        return False

# READ ALL ROWS IN WORKSHEET AND TRANSFORM INTO JSON
def all_data_to_json(worksheet, filename, sheetname):
    max_row = worksheet.max_row
    max_column = worksheet.max_column

    header = []
    for col in worksheet.iter_rows(min_row=1, max_col=max_column, max_row=1):
        for cell in col:
            header.append(cell.value)

    data = []
    for row in worksheet.iter_rows(min_row=2, max_col=max_column, max_row=max_row):
        item = {}
        for cell in row: 
            try:
                if("Date" in header[column_index_from_string(cell.column)-1] or "Data" in header[column_index_from_string(cell.column)-1]):
                    item["Date"] = str(cell.value)
                    continue
            except:
                x = 0
            try:
                item[header[column_index_from_string(cell.column)-1].encode("utf-8")] = cell.value.encode("utf-8")
            except:
                try:
                    item[header[column_index_from_string(cell.column)-1].encode("utf-8")] = cell.value
                except:
                    try:
                        item[header[column_index_from_string(cell.column)-1]] = cell.value.encode("utf-8")
                    except:
                        item['None'] = cell.value
        data.append(item)
    
    json_data = {
        "FILE_NAME" : filename + "_" + sheetname,
        "HEADER" : header,
        "DATA_INFOS" : data,
        "DATA_NUMBER" : max_row - 1
    }

    if(insert_data_db(json_data)):
        return json_data["FILE_NAME"]
    

def identifier(data, word):
    for line in data["DATA_INFOS"]:
        for string in line:
            if(string == word):
                return True
        return False

@app.route("/api/")
def index():
    return "hello world"

@app.route("/api/upload", methods=["POST"])
def upload():
    results = []

    for upload in request.files.getlist("file"):
        # try:
        
        filename = upload.filename.split(".")
        filename = filename[0]

        wb = load_workbook(filename=upload)#, read_only=True)
        sheets = wb.sheetnames
        for sheet in sheets:
            ws = wb[sheet]
            result = all_data_to_json(ws, filename, sheet)
            if result != None:
                results.append(result)
        # except:
        #     print "erro"

    return jsonify(results)

@app.route("/api/download", methods=["GET"])
def download():
    try:
        converted = mongo.db.converted
        json_file = converted.find_one({'FILE_NAME': request.args.get('filename')})

        # delete id form json file
        del json_file["_id"]
        return jsonify(json_file)
    except:
        return jsonify({})

@app.route("/api/search", methods=["GET"])
def search():
    converted = mongo.db.converted
    # suggestion = converted.find_one({'HEADER': request.args.get('value')})
    # suggestion = converted.find({'HEADER': "a"}).limit(10)
    suggestions = []
    try:
        for data in converted.find():
            for header in data["HEADER"]:
                if None != header and len(suggestions) < 10:
                    if request.args.get('value') in header and header not in suggestions:
                        suggestions.append(header)
                if len(suggestions) == 10:
                    return jsonify(suggestions)
                    
        return jsonify(suggestions) 
    except:
        return jsonify(suggestions)

@app.route('/searchbackup', methods=['GET', 'POST'])
def searchbackup():
    if request.method == 'POST':
        word = request.form['word']

        files = os.listdir(DIR_JSON)

        result = []
        for file in files:
            if(".json" in file and not "result.json" in file):            
                data = json.load(open(DIR_JSON + file))
                
                if(identifier(data, word)):
                    result.append(file)
        if (len(result) > 0 ):
            with open('{}result.json'.format(DIR_JSON), 'w') as json_result:
                    result_data = json.load(open(DIR_JSON + result.pop()))

                    for file in result:
                        data = json.load(open(DIR_JSON + file))
                        row_data = []
                        for row in data["DATA_INFOS"]:#["identificador"]
                            row_data.append(row)
                            # d.update(result_data)
                        

                        for row in row_data:
                            for r in result_data["DATA_INFOS"]:
                                if (row["identificador"] == r["identificador"]):
                                    r.update(row)
                                    break

                        count = 0
                        for i in result_data["DATA_INFOS"]:
                            count = count + 1

                        result_data["DATA_NUMBER"] = count

                    json.dump(result_data, json_result, indent = 4)
                    json_result.close()
        else:
            return render_template('search.html', result={})            

        data = json.load(open(DIR_JSON + "result.json"))

        return render_template('search.html', result=data) 
    else:
        return render_template('search.html', result={})

# @app.route('/download')
# def download():

#     files = []
#     directory = os.listdir(DIR_JSON)
    
#     for file in directory:
#         if(".json" in file):
#             files.append(file)

#     return render_template('download.html', files=files)